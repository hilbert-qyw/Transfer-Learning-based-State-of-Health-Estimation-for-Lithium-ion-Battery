# coding = utf-8
"""
作者   : Hilbert
时间   :2022/7/12 17:14
"""
import sys
import os
from warnings import simplefilter

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]  # 上一级目录
PathProject = os.path.split(rootPath)[0]
sys.path.append(rootPath)
sys.path.append(PathProject)
simplefilter(action='ignore', category=Warning)
simplefilter(action='ignore', category=FutureWarning)



import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy.io as sio
import datetime
from openpyxl import Workbook, load_workbook
from time import time
import xlrd
from tqdm import tqdm
import pickle


def mkdir(path):
    """
    mkdir of the path
    :param input: string of the path
    return: boolean
    """
    path = path.strip()
    path = path.rstrip('\\')
    isExists = os.path.exists(path)
    if not isExists:
        os.makedirs(path)
        print(path + ' is created!')
        return True
    else:
        print(path+' already exists!')
        return False


def mat2pk(input_path: str, saving_path: str, file_name: str):
    raw_data = load_workbook(os.path.join(input_path, file_name + '.xlsx'), read_only=True)
    sheet_name = raw_data.sheetnames
    battery_data = {}

    for name in tqdm(sheet_name, desc='Waiting...'):
        # with open(os.path.join(input_path, file_name + '.xlsx'), 'rb') as f:
        #     content = f.read()
        raw_data = pd.read_excel(os.path.join(input_path, file_name + '.xlsx'), sheet_name=name, engine='openpyxl')
        raw_data = raw_data.to_dict(orient='series')
        if len(raw_data) > 1:
            battery_data[name] = pd.DataFrame(raw_data)

    with open(os.path.join(saving_path, file_name +'.pk'), 'wb') as f:
        pickle.dump(battery_data, f)


class BatteryDataSegment(object):
    def __init__(self, input_path: str, saving_path: str, file_name: str, data_range: dict):
        self.input_path = input_path
        self.saving_path = saving_path
        self.name = file_name
        self.dataset = self.load_data()
        self.data_range = data_range
        self.font = {'family': 'Times New Roman', 'weight': 'normal', 'size': 32}

    """
    ['1st Charge1','1st Discharge2','2nd Charge3','50 Partial Cycles4','50 Partial Cycles5','50 Partial Cycles6',
    '50 Partial Cycles7','50 Partial Cycles8', '50 Partial Cycles9', '50 Partial Cycles10', '50 Partial Cycles11',
    '50 Partial Cycles12','50 Partial Cycles13', '50 Partial Cycles14', '50 Partial Cycles15', '50 Partial Cycles16',
    '50 Partial Cycles17', '50 Partial Cycles18', '50 Partial Cycles19', '50 Partial Cycles20', '50 Partial Cycles21',
    '50 Partial Cycles22', '50 Partial Cycles23', '50 Partial Cycles24', '50 Partial Cycles25', '50 Partial Cycles26',
    'Single Cycle29', '50 Partial Cycles30', '50 Partial Cycles31', '50 Partial Cycles32', '50 Partial Cycles33',
    '50 Partial Cycles34', '50 Partial Cycles35', '50 Partial Cycles36', '50 Partial Cycles37', '100 Partial Cycles38']
    
    data index:
    ['Time_sec', 'Date_Time', 'Step', 'Cycle', 'Current_Amp', 'Voltage_Volt',
     'Charge_Ah', 'Discharge_Ah']
    """

    def load_data(self) -> dict:
        print("=========Load data===============")
        with open(os.path.join(self.input_path, self.name + '.pk'), 'rb') as f:
            raw_data = pickle.load(f)
        print("Finish!")
        return raw_data

    def period_cycle(self, save_plot=True):
        path = self.saving_path + '/period_cycle/'
        mkdir(path)

        """
        test integrating
        
        """

        period_capacity = []
        for name, period_data in tqdm(self.dataset.items()):
            time = period_data['Time_sec'] / 3600
            current = period_data['Current_Amp']
            voltage = period_data['Voltage_Volt']
            charge = period_data['Charge_Ah']
            discharge = period_data['Discharge_Ah']

            # test integrating
            # print(np.trapz(current, time))

            self.plot_current(current, time, name='Period_cycle_data', path=path + name)
            self.plot_voltage(voltage, time, name='Period_cycle_data', path=path + name)
            self.plot_charge(charge, time, name='Period_cycle_data', path=path + name)

            # every 50 cycles discharge capacity
            if len(period_data) > 1e4 :
                period_capacity.append(period_data['Discharge_Ah'].iloc[-1])

        # plot capacity
        fig1 = plt.figure(figsize=(16, 12))
        ax1 = fig1.add_subplot(111)
        ax1.set_title('Discharge capacity', fontdict=self.font)
        ax1.set_xlabel('Cycles', fontdict=self.font)
        ax1.set_ylabel('Capacity', fontdict=self.font)
        ax1.plot(period_capacity)
        plt.tight_layout()
        plt.savefig(path + 'total_capacity.png')
        # plt.show()


    def each_cycle(self):
        path = self.saving_path + '/each_cycle/'
        mkdir(path)

        whole_curve = {}
        shallow_charge_curve = {}
        shallow_discharge_curve = {}
        calibrate_charge_curve = {}
        calibrate_discharge_curve = {}
        print(list(self.dataset.keys()))
        for name, period_data in tqdm(self.dataset.items()):
            if ('Partial' in name) or ('Full' in name):
                whole_curve[name] = []
                shallow_charge_curve[name] = []
                shallow_discharge_curve[name] = []
                calibrate_discharge_curve[name] = []
                calibrate_charge_curve[name] = []
                cycle_id = period_data['Cycle'].values
                cycle_sort = np.unique(cycle_id)
                for id in cycle_sort:

                    # whole curve
                    tmp_cycle = period_data.loc[period_data[period_data['Cycle']==id].index]
                    # print(np.unique(period_data['Step'].values))

                    # divide cycle
                    # print(np.unique(tmp_cycle['Step'].values))
                    # judge the effectiveness of the data
                    # if tmp_cycle.shape[0] < 10:
                    #     continue
                    # else:

                    # calculate capacity
                    time_cycle = tmp_cycle['Time_sec']
                    time_cycle = (time_cycle-time_cycle.iloc[0]) / 3600
                    current_cycle = tmp_cycle['Current_Amp']
                    each_step_capacity = []
                    for i in range(len(time_cycle)):
                        each_step_capacity.append(np.trapz(current_cycle[:i+1], time_cycle[:i+1]))
                    tmp_cycle['Capacity'] = each_step_capacity
                    whole_curve[name].append(tmp_cycle)

                    # shallow charge and discharge curve
                    # When the battery is during shallow charge and discharge, the step is 6, 7, 8, 9
                    # 6 --> shallow charge
                    # 7, 8, 9 --> shallow discharge
                    # shallow_charge = tmp_cycle.loc[tmp_cycle['Step']==self.data_range['partial_shallow_chg'][0]]
                    period_cycle_flag = name.split('Cycles')[1]

                    if int(period_cycle_flag) < 12:
                        shallow_charge = tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                                           True if (self.data_range[
                                                                                        'partial_shallow_chg'][0] < x) &
                                                                                   (x < self.data_range[
                                                                                       'partial_shallow_chg'][1])
                                                                           else False)]
                        shallow_discharge_curve[name].append(
                            tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                              True if (self.data_range['partial_shallow_dchg'][0] < x) &
                                                                      (x < self.data_range['partial_shallow_dchg'][1])
                                                              else False)])
                    else:
                        shallow_charge = tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                                           True if (6 < x) &
                                                                                   (x < 8)
                                                                           else False)]
                        shallow_discharge_curve[name].append(
                            tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                              True if (7 < x) &
                                                                      (x < 11)
                                                              else False)])
                    if len(shallow_charge) > 0:
                    # refresh capacity
                        time_cycle = shallow_charge['Time_sec']
                        time_cycle = (time_cycle - time_cycle.iloc[0]) / 3600
                        current_cycle = shallow_charge['Current_Amp']
                        each_step_capacity = []
                        for i in range(len(time_cycle)):
                            each_step_capacity.append(np.trapz(current_cycle[:i + 1], time_cycle[:i + 1]))
                        shallow_charge['Capacity'] = each_step_capacity

                        shallow_charge_curve[name].append(shallow_charge)


                    # fully charge and discharge curve
                    if int(period_cycle_flag) < 12:
                        # 13-15 fully discharge
                        fully_discharge = tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                                            True if (self.data_range['partial_fully_dchg'][0] < x) &
                                                                                    (x < self.data_range['partial_fully_dchg'][1])
                                                                            else False)]
                        # 16-17 fully charge
                        fully_charge = tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                                         True if (self.data_range['partial_fully_chg'][
                                                                                      0] < x) &
                                                                                 (x <
                                                                                  self.data_range['partial_fully_chg'][
                                                                                      1])
                                                                         else False)]

                    else:
                        # 13-15 fully discharge
                        fully_discharge = tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                                            True if (13 < x) &
                                                                                    (x < 17)
                                                                            else False)]
                        # 16-17 fully charge
                        fully_charge = tmp_cycle[tmp_cycle['Step'].apply(lambda x:
                                                                         True if (16 < x) &
                                                                                 (x < 19)
                                                                         else False)]

                    if len(fully_discharge) > 0:
                        calibrate_discharge_curve[name].append(fully_discharge)

                    if len(fully_charge) > 0:
                        # refresh capacity
                        time_cycle = fully_charge['Time_sec']
                        time_cycle = (time_cycle - time_cycle.iloc[0]) / 3600
                        current_cycle = fully_charge['Current_Amp']
                        each_step_capacity = []
                        for i in range(len(time_cycle)):
                            each_step_capacity.append(np.trapz(current_cycle[:i + 1], time_cycle[:i + 1]))
                        fully_charge['Capacity'] = each_step_capacity
                        if each_step_capacity[-1] > 1:
                            calibrate_charge_curve[name].append(fully_charge)

                # each cycle plot
                plot_time = np.concatenate([tmp_cycle['Time_sec'] / 3600 for tmp_cycle in whole_curve[name]])
                plot_capacity = np.concatenate([tmp_cycle['Capacity'] for tmp_cycle in whole_curve[name]])
                self.plot_charge(plot_capacity, plot_time, name='Each_cycle_data', path=path + name)

            elif 'Single' in name:
                # state: 1-12
                calibrate_discharge_curve[name] = []
                calibrate_charge_curve[name] = []
                # fully charge and discharge curve
                # 5-8 fully discharge
                period_data = self.dataset[name]
                fully_discharge = period_data[period_data['Step'].apply(lambda x:
                                                                        True if (self.data_range['single_dchg'][0] < x) &
                                                                                (x < self.data_range['single_dchg'][1])
                                                                        else False)]
                if len(fully_discharge) > 0:
                    calibrate_discharge_curve[name].append(fully_discharge)

                # 9-10 fully charge
                fully_charge = period_data[period_data['Step'].apply(lambda x:
                                                                     True if (self.data_range['single_chg'][0] < x) &
                                                                             (x < self.data_range['single_chg'][1])
                                                                     else False)]
                if len(fully_charge) > 0:
                    # refresh capacity
                    time_cycle = fully_charge['Time_sec']
                    time_cycle = (time_cycle - time_cycle.iloc[0]) / 3600
                    current_cycle = fully_charge['Current_Amp']
                    each_step_capacity = []
                    for i in range(len(time_cycle)):
                        each_step_capacity.append(np.trapz(current_cycle[:i + 1], time_cycle[:i + 1]))
                    fully_charge['Capacity'] = each_step_capacity

                    if each_step_capacity[-1] > 1:
                        calibrate_charge_curve[name].append(fully_charge)

            elif 'Discharge' in name:
                calibrate_discharge_curve[name] = []
                period_data = self.dataset[name]
                # state: 2-4
                fully_discharge = period_data[period_data['Step'].apply(lambda x:
                                                                        True if (self.data_range['discharge'][0] < x) &
                                                                                (x < self.data_range['discharge'][1])
                                                                        else False)]
                if len(fully_discharge) > 0:
                    calibrate_discharge_curve[name].append(fully_discharge)

            elif 'Charge' in name:
                calibrate_charge_curve[name] = []
                period_data = self.dataset[name]
                # state: 2, 3
                fully_charge = period_data[period_data['Step'].apply(lambda x:
                                                                     True if (self.data_range['charge'][0] < x) &
                                                                             (x < self.data_range['charge'][1]) else False)]
                if len(fully_charge) > 0:
                    time_cycle = fully_charge['Time_sec']
                    time_cycle = (time_cycle - time_cycle.iloc[0]) / 3600
                    current_cycle = fully_charge['Current_Amp']
                    each_step_capacity = []
                    for i in range(len(time_cycle)):
                        each_step_capacity.append(np.trapz(current_cycle[:i + 1], time_cycle[:i + 1]))
                    fully_charge['Capacity'] = each_step_capacity

                    if each_step_capacity[-1] > 1:
                        calibrate_charge_curve[name].append(fully_charge)

        # save shallow and fully charge curve
        charge_curve = [shallow_charge_curve, calibrate_charge_curve]
        discharge_curve = [shallow_discharge_curve, calibrate_discharge_curve]
        path = '/public/home/yuwen_hilbert/Program/CALCE_battery/battery_data/data_curve/'
        mkdir(path)
        with open(fr'{path}{self.name}_charge_curve.pk', 'wb') as f:
            pickle.dump(charge_curve, f)

        with open(fr'{path}{self.name}_discharge_curve.pk', 'wb') as f:
            pickle.dump(discharge_curve, f)

    def plot_current(self, current, time, name, path):
        fig1 = plt.figure(figsize=(16, 12))
        ax1 = fig1.add_subplot(111)
        ax1.set_title(name, fontdict=self.font)
        ax1.set_xlabel('Time (h)', fontdict=self.font)
        ax1.set_ylabel('Current (A)', fontdict=self.font)
        ax1.plot(time, current)
        plt.tight_layout()
        plt.savefig(path+'_current.png')
        # plt.show()
        # plt.close()

    def plot_voltage(self, current, time, name, path):
        fig2 = plt.figure(figsize=(16, 12))
        ax2 = fig2.add_subplot(111)
        ax2.set_title(name, fontdict=self.font)
        ax2.set_xlabel('Time (h)', fontdict=self.font)
        ax2.set_ylabel('Voltage (V)', fontdict=self.font)
        ax2.plot(time, current)
        plt.tight_layout()
        plt.savefig(path+'_voltage.png')
        # plt.show()
        # plt.close()

    def plot_charge(self, charge, time, name, path):
        fig3 = plt.figure(figsize=(16, 12))
        ax3 = fig3.add_subplot(111)
        ax3.set_title(name, fontdict=self.font)
        ax3.set_xlabel('Time (h)', fontdict=self.font)
        ax3.set_ylabel('Charge (Ah)', fontdict=self.font)
        ax3.plot(time, charge)
        plt.tight_layout()
        plt.savefig(path + '_capacity.png')
        # plt.show()
        # plt.close()


    def sensor_time(self, sensor_type):
        # voltage and current curve

        # delete value less than a certain value
        index = self.dataset['current_measured'][self.dataset['current_measured'].abs() < 0.01].index.tolist()
        self.dataset.drop(index=index, inplace=True)
        cycle = self.dataset['cycle'].unique()
        fig1 = plt.figure(figsize=(16, 9))
        ax1 = fig1.add_subplot(111)
        ax1.grid(alpha=0.4, linestyle=':')
        for idx in cycle:
            sensor_data = self.dataset.loc[self.dataset['cycle'][self.dataset['cycle'] == idx].index]
            ax1.plot(sensor_data['time'], sensor_data[sensor_type+'_measured'], lw=2)
        ax1.set_xlabel('Time', fontdict=self.font)
        ax1.set_ylabel(sensor_type, fontdict=self.font)
        ax1.set_title(sensor_type + ' changes with time', fontdict=self.font)
        plt.tight_layout()
        fig1.savefig(self.saving_path + sensor_type + ' changes with time.png')
        plt.show()


class BatteryDataAnalysis(object):
    def __init__(self, input_path: str, saving_path: str, file_name: str):
        self.input_path = input_path
        self.saving_path = saving_path
        self.name = file_name
        self.shallow_curve, self.fully_curve = self.load_data()
        self.font = {'family': 'Times New Roman', 'weight': 'normal', 'size': 32}

    def load_data(self):
        print("=========Load data===============")
        with open(os.path.join(self.input_path, self.name + '_charge_curve.pk'), 'rb') as f:
            raw_data = pickle.load(f)
        print("Finish!")
        return raw_data[0], raw_data[1]

    def fully_curve_plot(self, saving=True):

        # calibrate_total_capacity = {}
        # for period, fully_charge_data in self.fully_curve.items():
        #     if len(fully_charge_data) > 0:
        #         fully_charge_data = fully_charge_data[0]
        #         calibrate_total_capacity[period] = fully_charge_data['Capacity'].iloc[-1]

        calibrate_total_capacity = []
        calibrate_time = []
        calibrate_voltage = []
        calibrate_current = []
        calibrate_step_capacity = []
        for period, fully_charge_data in self.fully_curve.items():
            if len(fully_charge_data) > 0:
                fully_charge_data = fully_charge_data[0]
                calibrate_total_capacity.append(fully_charge_data['Capacity'].iloc[-1])
                calibrate_time.append((fully_charge_data['Time_sec']-fully_charge_data['Time_sec'].iloc[0])/3600)
                calibrate_current.append(fully_charge_data['Current_Amp'])
                calibrate_voltage.append(fully_charge_data['Voltage_Volt'])
                calibrate_step_capacity.append(fully_charge_data['Capacity'])

        # plot figure
        self.plot_curve(title='Capacity v.s. time', xlabel='Calibration time', ylabel='Capacity',
                        x_axis=[list(range(len(calibrate_total_capacity)))], y_axis=[calibrate_total_capacity],
                        name='Fully_charge_capacity_fade_curve')

        self.plot_curve(title='Current v.s. time', xlabel='Time (h)', ylabel='Current (A)',
                        x_axis=calibrate_time, y_axis=calibrate_current,
                        name='Fully_charge_current_curve')

        self.plot_curve(title='Voltage v.s. time', xlabel='Time (h)', ylabel='Voltage (V)',
                        x_axis=calibrate_time, y_axis=calibrate_voltage,
                        name='Fully_charge_voltage_curve')

        self.plot_curve(title='Capacity v.s. time', xlabel='Time (h)', ylabel='Capacity (Ah)',
                        x_axis=calibrate_time, y_axis=calibrate_step_capacity,
                        name='Fully_charge_capacity_curve')

    def shallow_curve_plot(self):
        whole_total_capacity = []
        for period, shallow_charge_data in self.shallow_curve.items():
            charge_total_capacity = []
            charge_time = []
            charge_voltage = []
            charge_current = []
            charge_step_capacity = []
            diff_time = []
            if len(shallow_charge_data) > 0:
                for cycle_data in shallow_charge_data:
                    charge_total_capacity.append(cycle_data['Capacity'].iloc[-1])
                    charge_time.append((cycle_data['Time_sec']-cycle_data['Time_sec'].iloc[0])/3600)
                    charge_current.append(cycle_data['Current_Amp'])
                    charge_voltage.append(cycle_data['Voltage_Volt'])
                    charge_step_capacity.append(cycle_data['Capacity'])

                    diff_time.append(cycle_data['Time_sec'].diff(periods=1).dropna())
                whole_total_capacity.append(charge_total_capacity)
                # plot every period figure
                # self.plot_curve(title='Capacity v.s. time', xlabel='Shallow charge time', ylabel='Capacity',
                #                 x_axis=[list(range(len(charge_total_capacity)))], y_axis=[charge_total_capacity],
                #                 name=fr'{period}_Shallow_charge_capacity_fade_curve')
                #
                # self.plot_curve(title='Current v.s. time', xlabel='Time (h)', ylabel='Current (A)',
                #                 x_axis=charge_time, y_axis=charge_current,
                #                 name=fr'{period}_Shallow_charge_current_curve')
                #
                # self.plot_curve(title='Voltage v.s. time', xlabel='Time (h)', ylabel='Voltage (V)',
                #                 x_axis=charge_time, y_axis=charge_voltage,
                #                 name=fr'{period}_Shallow_charge_voltage_curve')
                #
                # self.plot_curve(title='Capacity v.s. time', xlabel='Time (h)', ylabel='Capacity (Ah)',
                #                 x_axis=charge_time, y_axis=charge_step_capacity,
                #                 name=fr'{period}_Shallow_charge_capacity_curve')

        # The whole life capacity curve
        whole_life_capacity = [cycle_capacity for period_capacity in whole_total_capacity
                               for cycle_capacity in period_capacity]
        self.plot_curve(title='Whole life Capacity v.s. time', xlabel='Shallow charge time', ylabel='Capacity',
                        x_axis=[list(range(len(whole_life_capacity)))], y_axis=[whole_life_capacity],
                        name=fr'whole_life_shallow_charge_capacity_fade_curve')

        # sensor sample time distribution
        diff_time = np.concatenate([cycle_diff_time for cycle_diff_time in diff_time])
        # the histogram of sample time
        fig = plt.figure(figsize=(16, 12))
        ax = fig.add_subplot(111)
        ax.set_title('Sample span time', fontdict=self.font)
        ax.set_xlabel('span time', fontdict=self.font)
        ax.set_ylabel('Frequency', fontdict=self.font)
        ax.hist(diff_time, histtype='bar', rwidth=0.8)
        plt.tight_layout()
        plt.savefig(self.saving_path + '/Sensor sample time.png')



    def plot_curve(self, title, xlabel, ylabel, x_axis, y_axis, name):
        fig = plt.figure(figsize=(16, 12))
        ax = fig.add_subplot(111)
        ax.set_title(title, fontdict=self.font)
        ax.set_xlabel(xlabel, fontdict=self.font)
        ax.set_ylabel(ylabel, fontdict=self.font)
        for (x, y) in zip(x_axis, y_axis):
            ax.plot(x, y)
        plt.tight_layout()
        plt.savefig(self.saving_path + fr'/{name}.png')
        # plt.show()


class BatteryFullyDataAnalysis(object):
    def __init__(self, input_path: str, saving_path: str, file_name: str):
        self.input_path = input_path
        self.saving_path = saving_path
        self.name = file_name
        self.shallow_curve, self.fully_curve = self.load_data()
        self.font = {'family': 'Times New Roman', 'weight': 'normal', 'size': 32}

    def load_data(self):
        print("=========Load data===============")
        with open(os.path.join(self.input_path, self.name + '_charge_curve.pk'), 'rb') as f:
            raw_data = pickle.load(f)
        print("Finish!")
        return raw_data[0], raw_data[1]

    def fully_charge_plot(self):
        whole_total_capacity = []
        for period, fully_charge_data in self.fully_curve.items():
            charge_total_capacity = []
            charge_time = []
            charge_voltage = []
            charge_current = []
            charge_step_capacity = []
            diff_time = []
            if len(fully_charge_data) > 0:
                for cycle_data in fully_charge_data:
                    if cycle_data['Capacity'].iloc[-1] > 1:
                        charge_total_capacity.append(cycle_data['Capacity'].iloc[-1])
                        charge_time.append((cycle_data['Time_sec']-cycle_data['Time_sec'].iloc[0])/3600)
                        charge_current.append(cycle_data['Current_Amp'])
                        charge_voltage.append(cycle_data['Voltage_Volt'])
                        charge_step_capacity.append(cycle_data['Capacity'])

                    diff_time.append(cycle_data['Time_sec'].diff(periods=1).dropna())
                whole_total_capacity.append(charge_total_capacity)
                # plot every period figure
                self.plot_curve(title='Capacity v.s. time', xlabel='Fully charge time', ylabel='Capacity',
                                x_axis=[list(range(len(charge_total_capacity)))], y_axis=[charge_total_capacity],
                                name=fr'{period}_Fully_charge_capacity_fade_curve')

                self.plot_curve(title='Current v.s. time', xlabel='Time (h)', ylabel='Current (A)',
                                x_axis=charge_time, y_axis=charge_current,
                                name=fr'{period}_Fully_charge_current_curve')

                self.plot_curve(title='Voltage v.s. time', xlabel='Time (h)', ylabel='Voltage (V)',
                                x_axis=charge_time, y_axis=charge_voltage,
                                name=fr'{period}_Fully_charge_voltage_curve')

                self.plot_curve(title='Capacity v.s. time', xlabel='Time (h)', ylabel='Capacity (Ah)',
                                x_axis=charge_time, y_axis=charge_step_capacity,
                                name=fr'{period}_Fully_charge_capacity_curve')

        # The whole life capacity curve
        whole_life_capacity = [cycle_capacity for period_capacity in whole_total_capacity
                               for cycle_capacity in period_capacity]
        self.plot_curve(title='Whole life Capacity v.s. time', xlabel='Fully charge time', ylabel='Capacity',
                        x_axis=[list(range(len(whole_life_capacity)))], y_axis=[whole_life_capacity],
                        name=fr'whole_life_fully_charge_capacity_fade_curve')

        # sensor sample time distribution
        diff_time = np.concatenate([cycle_diff_time for cycle_diff_time in diff_time])
        # the histogram of sample time
        fig = plt.figure(figsize=(16, 12))
        ax = fig.add_subplot(111)
        ax.set_title('Sample span time', fontdict=self.font)
        ax.set_xlabel('span time', fontdict=self.font)
        ax.set_ylabel('Frequency', fontdict=self.font)
        ax.hist(diff_time, histtype='bar', rwidth=0.8)
        plt.tight_layout()
        plt.savefig(self.saving_path + '/Sensor sample time.png')



    def plot_curve(self, title, xlabel, ylabel, x_axis, y_axis, name):
        fig = plt.figure(figsize=(16, 12))
        ax = fig.add_subplot(111)
        ax.set_title(title, fontdict=self.font)
        ax.set_xlabel(xlabel, fontdict=self.font)
        ax.set_ylabel(ylabel, fontdict=self.font)
        for (x, y) in zip(x_axis, y_axis):
            ax.plot(x, y)
        plt.tight_layout()
        plt.savefig(self.saving_path + fr'/{name}.png')
        # plt.show()